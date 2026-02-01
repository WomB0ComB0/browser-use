"""Extraction logic for Microsoft Excel files.

Handles .xlsx and .xlsm formats using the openpyxl library, extracting tabular
data into markdown format and providing structural metadata for AI processing.
"""

from __future__ import annotations

from pathlib import Path
from typing import TypedDict

from pipeline.extractors.base import BaseExtractor, ExtractedContent

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SheetInfo(TypedDict, total=False):
    """Deep structural information for a single Excel worksheet.

    Attributes:
        name: The display name of the sheet.
        row_count: Total number of rows found.
        column_count: Number of columns in the header row.
        headers: List of the first 20 column headers.
        data_types: Predicted data types for columns (numeric, date, text, etc.).
        sample_values: First 3 rows of data for preview.
    """
    name: str
    row_count: int
    column_count: int
    headers: list[str]
    data_types: list[str]
    sample_values: list[list[str]]


class ExcelStructure(TypedDict, total=False):
    """Complete structural manifest of an Excel workbook.

    Attributes:
        sheet_count: Total number of sheets in the workbook.
        sheet_names: Ordered list of all sheet names.
        sheets: List of detailed SheetInfo for each sheet.
        has_formulas: Whether the original file contained formulas (detected).
        has_charts: Whether the original file contained charts (detected).
    """
    sheet_count: int
    sheet_names: list[str]
    sheets: list[SheetInfo]
    has_formulas: bool
    has_charts: bool


class ExcelExtractor(BaseExtractor):
    """Handler for extracting text and structure from Excel files.

    Uses `openpyxl` to parse worksheets. For large files, it samples the content
    to avoid overwhelming LLM context limits while still providing a structural
    overview.
    """

    SUPPORTED_EXTENSIONS = [".xlsx", ".xlsm"]

    def supports(self, file_path: Path) -> bool:
        """Check if this extractor supports the given file."""
        return file_path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def extract(self, file_path: Path) -> ExtractedContent:
        """Extract content from an Excel file."""
        if not OPENPYXL_AVAILABLE:
            return self._create_unavailable_result(file_path)

        file_path = Path(file_path)
        metadata = self._get_file_metadata(file_path)

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            # Extract content from all sheets
            content_parts = []
            sheets_info: list[SheetInfo] = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_content, sheet_info = self._extract_sheet(sheet, sheet_name)
                content_parts.append(f"## Sheet: {sheet_name}\n\n{sheet_content}")
                sheets_info.append(sheet_info)

            content = "\n\n---\n\n".join(content_parts)

            # Build structure
            structure: ExcelStructure = {
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": list(workbook.sheetnames),
                "sheets": sheets_info,
                "has_formulas": False,  # data_only=True, so we don't see formulas
                "has_charts": False,
            }

            # Generate summary
            summary = self._generate_summary(structure)

            workbook.close()

            return ExtractedContent(
                content=content,
                summary=summary,
                file_path=file_path,
                file_type="Excel",
                file_size_bytes=metadata["file_size_bytes"],
                modified_time=metadata["modified_time"],
                structure=structure,
                metadata={
                    "sheet_count": structure["sheet_count"],
                    "sheet_names": structure["sheet_names"],
                },
            )

        except Exception as e:
            return ExtractedContent(
                content=f"Error extracting Excel file: {e}",
                summary=f"Failed to extract Excel content: {e}",
                file_path=file_path,
                file_type="Excel",
                file_size_bytes=metadata["file_size_bytes"],
                modified_time=metadata["modified_time"],
                structure=None,
                metadata={"error": str(e)},
            )

    def _extract_sheet(self, sheet, sheet_name: str) -> tuple[str, SheetInfo]:
        """Extract content and info from a single worksheet."""
        rows_data = self._get_rows_data(sheet)
        headers = rows_data[0] if rows_data else []

        content = self._format_sheet_markdown(rows_data)
        data_types = self._detect_column_types_from_rows(rows_data, len(headers))

        sheet_info: SheetInfo = {
            "name": sheet_name,
            "row_count": len(rows_data),
            "column_count": len(headers),
            "headers": headers[:20],
            "data_types": data_types,
            "sample_values": rows_data[1:4] if len(rows_data) > 1 else [],
        }

        return content, sheet_info

    def _get_rows_data(self, sheet) -> list[list[str]]:
        """Extract row data from a sheet, converting None to empty string."""
        rows_data = []
        for row in sheet.iter_rows(max_row=100, values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            rows_data.append(row_values)
        return rows_data

    def _format_sheet_markdown(self, rows_data: list[list[str]]) -> str:
        """Format sheet data as a markdown table."""
        if not rows_data:
            return ""

        content_lines = []
        # Header row
        header_row = rows_data[0]
        content_lines.append("| " + " | ".join(header_row[:10]) + " |")
        content_lines.append("| " + " | ".join(["---"] * min(len(header_row), 10)) + " |")

        # Data rows (first 20)
        for row in rows_data[1:21]:
            content_lines.append("| " + " | ".join(row[:10]) + " |")

        if len(rows_data) > 21:
            content_lines.append(f"\n*... and {len(rows_data) - 21} more rows*")

        return "\n".join(content_lines)

    def _detect_column_types_from_rows(self, rows_data: list[list[str]], col_count: int) -> list[str]:
        """Determine data types for the first few columns."""
        data_types = []
        if len(rows_data) <= 1:
            return data_types

        for col_idx in range(min(col_count, 10)):
            sample_values = []
            for row_idx in range(1, min(6, len(rows_data))):
                if col_idx < len(rows_data[row_idx]):
                    sample_values.append(rows_data[row_idx][col_idx])
            
            data_types.append(self._detect_column_type(sample_values))
        
        return data_types

    def _detect_column_type(self, values: list[str]) -> str:
        """Detect the predominant type of values in a column."""
        if not values:
            return "unknown"

        numeric_count = 0
        date_count = 0

        for val in values:
            if not val or val == "None":
                continue
            try:
                float(val.replace(",", ""))
                numeric_count += 1
            except ValueError:
                if any(sep in val for sep in ["-", "/", ":"]):
                    date_count += 1

        total = len([v for v in values if v and v != "None"])
        if total == 0:
            return "empty"
        if numeric_count / total > 0.8:
            return "numeric"
        if date_count / total > 0.5:
            return "date"
        return "text"

    def _generate_summary(self, structure: ExcelStructure) -> str:
        """Generate a summary of the Excel file."""
        sheet_count = structure["sheet_count"]
        sheet_names = structure["sheet_names"]

        summary_parts = [f"Excel workbook with {sheet_count} sheet(s)"]

        if sheet_names:
            summary_parts.append(f"Sheets: {', '.join(sheet_names[:5])}")
            if len(sheet_names) > 5:
                summary_parts.append(f"and {len(sheet_names) - 5} more")

        # Add total row/column counts
        total_rows = sum(s.get("row_count", 0) for s in structure.get("sheets", []))
        summary_parts.append(f"Total rows across sheets: {total_rows:,}")

        return ". ".join(summary_parts) + "."

    def _create_unavailable_result(self, file_path: Path) -> ExtractedContent:
        """Create result when openpyxl is not available."""
        metadata = self._get_file_metadata(file_path)
        return ExtractedContent(
            content="Excel extraction requires openpyxl. Install with: pip install openpyxl",
            summary="Excel extraction unavailable - openpyxl not installed",
            file_path=file_path,
            file_type="Excel",
            file_size_bytes=metadata["file_size_bytes"],
            modified_time=metadata["modified_time"],
            structure=None,
            metadata={"error": "openpyxl not installed"},
        )
